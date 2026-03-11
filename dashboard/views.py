from django.contrib.auth.decorators import login_required, user_passes_test
from store.models import SubCategory, Feedback
from store.forms import ProductForm
from store.models import Product, Category
from store.forms import CategoryForm
from django.contrib.auth.models import User
from store.forms import SubCategoryForm
from django.contrib import messages
import json
from store.models import Customer, Invoice, OrderSummary, ShippingAddress

from store.models import Order, OrderItem
from django.utils import timezone
from datetime import datetime

from datetime import timedelta
from collections import defaultdict
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from store.forms import CustomerForm
from store.forms import OrderSummaryForm
from store.forms import InvoiceForm
from store.forms import ShippingAddressForm
from django.db.models.functions import ExtractDay, ExtractMonth, TruncMonth, TruncYear
from django.db.models import Sum, Count, F, DecimalField, ExpressionWrapper
from store.forms import UserEditForm
from django.utils.dateparse import parse_datetime
from django.http import HttpResponse
import openpyxl
import csv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

def is_admin(user):
    return user.is_staff or user.is_superuser


@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def dashboard_edit_product(request, pk):
    product = get_object_or_404(Product, pk=pk)

    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            product = form.save(commit=False)

            # 🔥 GET COLORS FROM HIDDEN INPUT
            colors_json = request.POST.get("colors", "[]")

            try:
                product.colors = json.loads(colors_json)  # convert JSON string → Python list
            except json.JSONDecodeError:
                product.colors = []

            product.save()
            form.save_m2m()

            return redirect('dashboard:dashboard_products')
    else:
        form = ProductForm(instance=product)

    return render(request, 'dashboard/product_form.html', {
        'form': form,
        'title': 'Edit Product',
        'existing_colors': json.dumps(product.colors or [])

    })


# admin/staff check
def is_admin(user):
    return user.is_staff or user.is_superuser


@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def dashboard_products(request):
    products = Product.objects.select_related('subcategory')

    for product in products:
        if not product.sizes:
            product.size_list = []

        # If sizes is already a list (best case)
        elif isinstance(product.sizes, list):
            product.size_list = product.sizes

        # If sizes is a JSON string like '["L", "XL"]'
        elif isinstance(product.sizes, str):
            try:
                product.size_list = json.loads(product.sizes)
            except json.JSONDecodeError:
                product.size_list = []

        else:
            product.size_list = []

    return render(request, 'dashboard/products.html', {
        'products': products
    })

@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def dashboard_users(request):
    return render(request, 'dashboard/users.html')

@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def dashboard_add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)

            colors_json = request.POST.get("colors", "[]")
            try:
                product.colors = json.loads(colors_json)
            except json.JSONDecodeError:
                product.colors = []

            product.save()
            form.save_m2m()
            return redirect('dashboard:dashboard_products')

    else:
        form = ProductForm()

    return render(request, 'dashboard/product_form.html', {
        'form': form,
        'title': 'Add Product',
        'existing_colors': '[]'  # 🔥 ADD THIS LINE
    })


@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def dashboard_subcategories(request):
    subcategories = SubCategory.objects.all()  # Fetch all subcategories
    return render(request, 'dashboard/subcategories.html', {
        'subcategories': subcategories
    })

@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def dashboard_categories(request):
    categories = Category.objects.all()  # Fetch all categories
    return render(request, 'dashboard/categories.html', {
        'categories': categories
    })


# EDIT product
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            return redirect('product_list')
    else:
        form = ProductForm(instance=product)
    return render(request, 'dashboard/product_form.html', {'form': form, 'title': 'Edit Product'})


@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def dashboard_delete_product(request, pk):
    product = get_object_or_404(Product, pk=pk)

    if request.method == "POST":
        product.delete()
        return redirect('dashboard:dashboard_products')

    return render(request, "dashboard/product_confirm_delete.html", {
        "product": product
    })

@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def dashboard_add_category(request):
    form = CategoryForm()

    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('dashboard:dashboard_categories')

    return render(request, 'dashboard/category_form.html', {
        'form': form,
        'title': 'Add Category'
    })
def dashboard_edit_category(request, pk):
    category = Category.objects.get(pk=pk)
    form = CategoryForm(instance=category)

    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('dashboard:dashboard_categories')

    return render(request, 'dashboard/category_form.html', {'form': form})



@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def dashboard_home(request):
    context = {
        'total_products': Product.objects.count(),
        'total_categories': Category.objects.count(),
        'total_subcategories': SubCategory.objects.count(),
        'total_users': User.objects.count(),
    }
    return render(request, 'dashboard/home.html', context)


@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def add_subcategory(request):
    if request.method == 'POST':
        form = SubCategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('dashboard:dashboard_subcategories')
    else:
        form = SubCategoryForm()

    return render(request, 'dashboard/subcategory_form.html', {
        'form': form
    })


def dashboard_orders(request):
    query = request.GET.get('q')

    orders = Order.objects.all().order_by('-date_ordered')

    if query:
        orders = orders.filter(
            Q(id__icontains=query) |
            Q(customer__name__icontains=query)
        )

    context = {
        'orders': orders
    }
    return render(request, 'dashboard/orders.html', context)


@login_required
@user_passes_test(is_admin)
def users_list(request):
    users = User.objects.all()
    return render(request, 'dashboard/users.html', {'users': users})

@login_required
@user_passes_test(is_admin)
def edit_subcategory(request, pk):
    subcategory = get_object_or_404(SubCategory, pk=pk)

    if request.method == 'POST':
        form = SubCategoryForm(request.POST, request.FILES, instance=subcategory)
        if form.is_valid():
            form.save()
            return redirect('dashboard:dashboard_subcategories')
    else:
        form = SubCategoryForm(instance=subcategory)

    return render(request, 'dashboard/subcategory_form.html', {'form': form})

@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def dashboard_delete_category(request, pk):
    category = get_object_or_404(Category, pk=pk)

    if request.method == 'POST':
        category.delete()
        return redirect('dashboard:dashboard_categories')

    return render(request, 'dashboard/category_confirm_delete.html', {
        'category': category
    })


@login_required
@user_passes_test(is_admin)
def delete_subcategory(request, pk):
    subcategory = get_object_or_404(SubCategory, pk=pk)

    if request.method == "POST":
        subcategory.delete()
        messages.success(request, "Subcategory deleted successfully.")
        return redirect("dashboard:dashboard_subcategories")

    return render(request, "dashboard/delete_subcategory.html", {"subcategory": subcategory})

@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def dashboard_customers(request):
    customers = Customer.objects.select_related('user').all()

    return render(request, 'dashboard/customers.html', {
        'customers': customers
    })

@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def dashboard_shipping_addresses(request):
    addresses = ShippingAddress.objects.select_related('customer', 'order').all()

    return render(request, 'dashboard/shipping_addresses.html', {
        'addresses': addresses
    })

@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def dashboard_order_summaries(request):
    summaries = OrderSummary.objects.select_related('order').all()

    return render(request, 'dashboard/order_summaries.html', {
        'summaries': summaries
    })

@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def dashboard_invoices(request):
    invoices = Invoice.objects.select_related('order').all()

    return render(request, 'dashboard/invoices.html', {
        'invoices': invoices
    })

def get_sales_data(request):
    from django.utils import timezone
    from datetime import timedelta
    from django.db.models import Sum, F, DecimalField, ExpressionWrapper
    from collections import defaultdict
    from store.models import Order, OrderItem

    # 📊 BASIC STATS
    total_orders = Order.objects.count()
    paid_orders = Order.objects.filter(complete=True).count()
    unpaid_orders = Order.objects.filter(complete=False).count()

    total_revenue = OrderItem.objects.filter(
        order__complete=True
    ).aggregate(
        total=Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    )['total'] or 0

    # 📅 SALES THIS MONTH
    today = timezone.now()
    first_day_of_month = today.replace(day=1)

    monthly_revenue = OrderItem.objects.filter(
        order__complete=True,
        order__date_ordered__gte=first_day_of_month
    ).aggregate(
        total=Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    )['total'] or 0

    # 📈 LAST 7 DAYS SALES
    today = timezone.now().date()
    seven_days_ago = today - timedelta(days=6)

    sales_data = OrderItem.objects.filter(
        order__complete=True,
        order__date_ordered__date__gte=seven_days_ago
    ).annotate(
        day=F('order__date_ordered__date')
    ).values('day').annotate(
        revenue=Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    ).order_by('day')

    sales_dict = defaultdict(int)
    for entry in sales_data:
        sales_dict[entry['day']] = float(entry['revenue'])

    dates = []
    revenues = []

    return {
        'total_orders': total_orders,
        'paid_orders': paid_orders,
        'unpaid_orders': unpaid_orders,
        'total_revenue': total_revenue,
        'monthly_revenue': monthly_revenue,
        'sales_dates': dates,
        'sales_revenues': revenues,
    }

@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def reports_dashboard(request):
    sales_data = get_sales_data(request)

    # ✅ Your model uses date_ordered, not created_at
    recent_orders = Order.objects.select_related('customer__user').order_by('-date_ordered')[:10]

    # 🏆 Top Selling Products
    top_products = OrderItem.objects.filter(order__complete=True) \
        .values('product__id', 'product__name', 'product__price') \
        .annotate(
        total_sold=Sum('quantity'),
        revenue=Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    ).order_by('-total_sold')[:10]

    # 📊 DATA FOR TOP SELLING PRODUCTS CHART
    product_names = [item['product__name'] for item in top_products]
    product_sales = [item['total_sold'] for item in top_products]

    # 📆 SALES THIS MONTH
    today = timezone.now()
    first_day_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)


    # 📈 SALES OVER LAST 7 DAYS
    today = timezone.now().date()
    seven_days_ago = today - timedelta(days=6)

    sales_queryset = OrderItem.objects.filter(
        order__complete=True,
        order__date_ordered__date__gte=seven_days_ago
    ).annotate(
        day=F('order__date_ordered__date')
    ).values('day').annotate(
        revenue=Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    ).order_by('day')

    # Fill missing days with 0 revenue
    sales_dict = defaultdict(int)
    for entry in sales_queryset:
        sales_dict[entry['day']] = float(entry['revenue'])

    # dates = []
    # revenues = []

    total_orders = sales_data['total_orders']
    paid_orders = sales_data['paid_orders']
    unpaid_orders = sales_data['unpaid_orders']
    total_revenue = sales_data['total_revenue']
    monthly_revenue = sales_data['monthly_revenue']

    dates = sales_data['sales_dates']
    revenues = sales_data['sales_revenues']

    # 🥧 SALES BY CATEGORY
    category_data = OrderItem.objects.filter(order__complete=True) \
        .values('product__subcategory__category__name') \
        .annotate(
        revenue=Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    ).order_by('-revenue')

    category_labels = [item['product__subcategory__category__name'] for item in category_data]
    category_revenues = [float(item['revenue'] or 0) for item in category_data]


    # 📅 MONTHLY SALES COMPARISON (Last 6 Months)
    today = timezone.now()
    six_months_ago = today - timedelta(days=180)

    monthly_data = OrderItem.objects.filter(
        order__complete=True,
        order__date_ordered__gte=six_months_ago
    ).annotate(
        month=TruncMonth('order__date_ordered')
    ).values('month').annotate(
        revenue=Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    ).order_by('month')

    month_labels = [entry['month'].strftime('%b %Y') for entry in monthly_data]
    month_revenues = [float(entry['revenue']) for entry in monthly_data]

    # 📆 YEARLY SALES COMPARISON
    yearly_data = OrderItem.objects.filter(order__complete=True) \
        .annotate(
        year=TruncYear('order__date_ordered')
    ).values('year').annotate(
        revenue=Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    ).order_by('year')

    year_labels = [entry['year'].strftime('%Y') for entry in yearly_data]
    year_revenues = [float(entry['revenue']) for entry in yearly_data]

    # 📅 MONTHLY SALES (JAN–DEC FOR CURRENT YEAR)
    selected_year = request.GET.get('year')

    if selected_year:
        selected_year = int(selected_year)
    else:
        selected_year = timezone.now().year

    monthly_sales = OrderItem.objects.filter(
        order__complete=True,
        order__date_ordered__year=selected_year
    ).annotate(
        month=ExtractMonth('order__date_ordered')
    ).values('month').annotate(
        revenue=Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    )

    # Convert to dictionary {month_number: revenue}
    sales_dict = {entry['month']: float(entry['revenue']) for entry in monthly_sales}

    month_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                    'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    month_revenues = [sales_dict.get(i, 0) for i in range(1, 13)]

    # 📈 REVENUE vs ORDERS TREND (Last 7 Days)
    today = timezone.now().date()
    seven_days_ago = today - timedelta(days=6)

    orders_data = Order.objects.filter(
        complete=True,
        date_ordered__date__gte=seven_days_ago
    ).values('date_ordered__date').annotate(
        order_count=Count('id')
    ).order_by('date_ordered__date')

    revenue_data = OrderItem.objects.filter(
        order__complete=True,
        order__date_ordered__date__gte=seven_days_ago
    ).annotate(
        day=F('order__date_ordered__date')
    ).values('day').annotate(
        revenue=Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    )

    orders_dict = {entry['date_ordered__date']: entry['order_count'] for entry in orders_data}
    revenue_dict = {entry['day']: float(entry['revenue']) for entry in revenue_data}

    trend_dates = []
    trend_orders = []
    trend_revenue = []

    for i in range(7):
        day = seven_days_ago + timedelta(days=i)
        trend_dates.append(day.strftime("%d %b"))
        trend_orders.append(orders_dict.get(day, 0))
        trend_revenue.append(revenue_dict.get(day, 0))

    # 💰 REVENUE BY PRICE RANGE
    price_ranges = {
        "Under ₹500": OrderItem.objects.filter(
            order__complete=True,
            product__price__lt=500
        ),
        "₹500 - ₹1000": OrderItem.objects.filter(
            order__complete=True,
            product__price__gte=500,
            product__price__lte=1000
        ),
        "Above ₹1000": OrderItem.objects.filter(
            order__complete=True,
            product__price__gt=1000
        ),
    }

    range_labels = []
    range_revenues = []

    for label, queryset in price_ranges.items():
        revenue = queryset.aggregate(
            total=Sum(
                ExpressionWrapper(
                    F('product__price') * F('quantity'),
                    output_field=DecimalField()
                )
            )
        )['total'] or 0

        range_labels.append(label)
        range_revenues.append(float(revenue))

    # 💳 AVERAGE ORDER VALUE (AOV) TREND – Last 7 Days
    today = timezone.now().date()
    seven_days_ago = today - timedelta(days=6)

    # Orders per day
    orders_per_day = Order.objects.filter(
        complete=True,
        date_ordered__date__gte=seven_days_ago
    ).values('date_ordered__date').annotate(
        count=Count('id')
    )

    # Revenue per day
    revenue_per_day = OrderItem.objects.filter(
        order__complete=True,
        order__date_ordered__date__gte=seven_days_ago
    ).annotate(
        day=F('order__date_ordered__date')
    ).values('day').annotate(
        revenue=Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    )

    orders_dict = {entry['date_ordered__date']: entry['count'] for entry in orders_per_day}
    revenue_dict = {entry['day']: float(entry['revenue']) for entry in revenue_per_day}

    aov_dates = []
    aov_values = []

    for i in range(7):
        day = seven_days_ago + timedelta(days=i)
        orders = orders_dict.get(day, 0)
        revenue = revenue_dict.get(day, 0)

        aov = revenue / orders if orders > 0 else 0

        aov_dates.append(day.strftime("%d %b"))
        aov_values.append(round(aov, 2))

    # 💎 CUSTOMER LIFETIME VALUE (CLV)

    customer_clv = OrderItem.objects.filter(order__complete=True, order__customer__user__isnull=False) \
        .values(
        'order__customer__id',
        'order__customer__user__username'
    ).annotate(
        total_spent=Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        ),
        total_orders=Count('order', distinct=True)
    ).order_by('-total_spent')[:10]  # Top 10 customers

    # 📊 CLV CHART DATA (Top 5 Customers)
    top_customers = list(customer_clv[:5])

    clv_labels = [c['order__customer__user__username'] for c in top_customers]
    clv_values = [float(c['total_spent']) for c in top_customers]

    # 📅 DAY-WISE REPORT (SELECT MONTH)
    selected_month = request.GET.get('month')
    selected_year = request.GET.get('year')

    import calendar

    # ✅ Default values (VERY IMPORTANT)
    selected_month_name = ""

    if selected_month:
        selected_month_name = calendar.month_name[int(selected_month)]

    day_labels = []
    day_orders = []
    day_revenues = []

    if selected_month and selected_year:
        selected_month = int(selected_month)
        selected_year = int(selected_year)

        # Get total days in month
        import calendar
        total_days = calendar.monthrange(selected_year, selected_month)[1]

        # Orders per day
        orders_data = Order.objects.filter(
            complete=True,
            date_ordered__year=selected_year,
            date_ordered__month=selected_month
        ).annotate(
            day=ExtractDay('date_ordered')
        ).values('day').annotate(
            total_orders=Count('id')
        )

        # Revenue per day
        revenue_data = OrderItem.objects.filter(
            order__complete=True,
            order__date_ordered__year=selected_year,
            order__date_ordered__month=selected_month
        ).annotate(
            day=ExtractDay('order__date_ordered')
        ).values('day').annotate(
            revenue=Sum(
                ExpressionWrapper(
                    F('product__price') * F('quantity'),
                    output_field=DecimalField()
                )
            )
        )

        # Convert to dictionary
        orders_dict = {item['day']: item['total_orders'] for item in orders_data}
        revenue_dict = {item['day']: float(item['revenue']) for item in revenue_data}

        # Fill ALL days (1 → 31)
        for day in range(1, total_days + 1):
            day_labels.append(str(day))
            day_orders.append(orders_dict.get(day, 0))
            day_revenues.append(revenue_dict.get(day, 0))

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    range_type = request.GET.get('range')

    today = timezone.now().date()

    if range_type == 'today':
        from_date = today
        to_date = today

    elif range_type == 'week':
        from_date = today - timedelta(days=6)
        to_date = today

    elif range_type == 'month':
        from_date = today.replace(day=1)
        to_date = today

    # ----------------------------------------
    # FILTERED DATE LOGIC (CLEAN VERSION)
    # ----------------------------------------

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    range_type = request.GET.get('range')

    today = timezone.now().date()

    if range_type == 'today':
        from_date = today
        to_date = today

    elif range_type == 'week':
        from_date = today - timedelta(days=6)
        to_date = today

    elif range_type == 'month':
        from_date = today.replace(day=1)
        to_date = today

    # Convert string dates properly
    if from_date and isinstance(from_date, str):
        from_date = datetime.strptime(from_date, "%Y-%m-%d").date()

    if to_date and isinstance(to_date, str):
        to_date = datetime.strptime(to_date, "%Y-%m-%d").date()

    filtered_orders = Order.objects.filter(complete=True)
    filtered_order_items = OrderItem.objects.filter(order__complete=True)

    if from_date and to_date:
        filtered_orders = filtered_orders.filter(
            date_ordered__date__range=[from_date, to_date]
        )

        filtered_order_items = filtered_order_items.filter(
            order__date_ordered__date__range=[from_date, to_date]
        )

    filtered_total_orders = filtered_orders.count()

    filtered_revenue = filtered_order_items.filter(
        order__complete=True
    ).aggregate(
        total=Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    )['total'] or 0

    # ----------------------------------------
    # FILTERED CHART DATA
    # ----------------------------------------

    filtered_sales_data = filtered_order_items.annotate(
        day=F('order__date_ordered__date')
    ).values('day').annotate(
        revenue=Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    ).order_by('day')

    sales_dict = defaultdict(float)

    for entry in filtered_sales_data:
        sales_dict[entry['day']] = float(entry['revenue'] or 0)

    filtered_dates = []
    filtered_revenues = []

    if from_date and to_date:
        current_day = from_date

        while current_day <= to_date:
            filtered_dates.append(current_day.strftime("%d %b"))
            filtered_revenues.append(sales_dict.get(current_day, 0))
            current_day += timedelta(days=1)



    user_report = OrderItem.objects.filter(
        order__complete=True,
        order__customer__user__isnull=False
    ).values(
        'order__customer__user__username'
    ).annotate(
        total_orders=Count('order', distinct=True),
        total_spent=Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    ).order_by('-total_spent')

    user_names = [u['order__customer__user__username'] for u in user_report]
    user_spending = [float(u['total_spent']) for u in user_report]

    context = {
        'total_orders': total_orders,
        'paid_orders': paid_orders,
        'unpaid_orders': unpaid_orders,
        'total_revenue': total_revenue,
        'monthly_revenue': monthly_revenue,
        'recent_orders': recent_orders,
        'top_products': top_products,
        'sales_dates': json.dumps(dates),
        'sales_revenues': json.dumps(revenues),
        'chart_product_names': json.dumps(product_names),
        'chart_product_sales': json.dumps(product_sales),
        'category_labels': json.dumps(category_labels),
        'category_revenues': json.dumps(category_revenues),
        'month_labels': json.dumps(month_labels),
        'month_revenues': json.dumps(month_revenues),
        'year_labels': json.dumps(year_labels),
        'year_revenues': json.dumps(year_revenues),
        'year_month_labels': json.dumps(month_labels),
        'year_month_revenues': json.dumps(month_revenues),
        'selected_year': selected_year,
        'available_years': range(timezone.now().year, timezone.now().year - 5, -1),
        'revenue_orders_dates': json.dumps(trend_dates),
        'revenue_orders_counts': json.dumps(trend_orders),
        'revenue_orders_revenue': json.dumps(trend_revenue),
        'price_range_labels': json.dumps(range_labels),
        'price_range_revenues': json.dumps(range_revenues),
        'aov_dates': json.dumps(aov_dates),
        'aov_values': json.dumps(aov_values),
        'customer_clv': customer_clv,
        'clv_labels': json.dumps(clv_labels),
        'clv_values': json.dumps(clv_values),
        'day_labels': json.dumps(day_labels),
        'day_orders': json.dumps(day_orders),
        'day_revenues': json.dumps(day_revenues),
        'selected_month': selected_month,
        'selected_month_name': selected_month_name,
        'filtered_total_orders': filtered_total_orders,
        'filtered_revenue': filtered_revenue,
        'filtered_dates': json.dumps(filtered_dates),
        'filtered_revenues': json.dumps(filtered_revenues),
        'selected_from_date': from_date,
        'selected_to_date': to_date,
        'user_report': user_report,
        'user_names': json.dumps(user_names),
        'user_spending': json.dumps(user_spending),
    }

    return render(request, 'dashboard/reports.html', context)


import csv
from django.http import HttpResponse
from store.models import Order   # ✅ correct import

def export_orders_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="orders.csv"'

    writer = csv.writer(response)
    writer.writerow(['Order ID', 'Customer', 'Date', 'Status', 'Total'])

    orders = Order.objects.all()

    for order in orders:
        customer = order.customer.user.username if order.customer and order.customer.user else "Guest"

        writer.writerow([
            order.id,
            customer,
            order.date_ordered.strftime("%d-%m-%Y"),
            "Paid" if order.complete else "Unpaid",
            order.get_cart_total
        ])

    return response

def export_sales_excel(request):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Header
    ws.append(["Order ID", "Customer", "Date", "Status", "Total"])

    # Header styling
    header_fill = PatternFill(start_color="0b2e66", end_color="0b2e66", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    orders = Order.objects.select_related('customer__user').all()

    for order in orders:

        customer = "Guest"
        if order.customer and order.customer.user:
            customer = order.customer.user.username

        ws.append([
            order.id,
            customer,
            order.date_ordered.strftime("%d-%m-%Y"),
            "Paid" if order.complete else "Unpaid",
            float(order.get_cart_total)
        ])


    table = Table(displayName="ReportTable", ref=f"A1:E{ws.max_row}")

    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)

    table.tableStyleInfo = style
    ws.add_table(table)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'

    wb.save(response)
    return response


def export_products_excel(request):
    from store.models import OrderItem

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products Report"

    # Header
    ws.append(["Product", "Units Sold", "Revenue"])

    # Header styling
    header_fill = PatternFill(start_color="0b2e66", end_color="0b2e66", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    products = OrderItem.objects.filter(order__complete=True) \
        .values('product__name') \
        .annotate(
            total_sold=Sum('quantity'),
            revenue=Sum(
                ExpressionWrapper(
                    F('product__price') * F('quantity'),
                    output_field=DecimalField()
                )
            )
        ).order_by('-total_sold')

    for p in products:
        ws.append([
            p['product__name'],
            p['total_sold'],
            float(p['revenue'] or 0)
        ])

    # Create table
    table = Table(displayName="ProductsTable", ref=f"A1:C{ws.max_row}")

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename="products_report.xlsx"'

    wb.save(response)
    return response

def export_products_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products_report.csv"'

    writer = csv.writer(response)
    writer.writerow(["Product", "Units Sold", "Revenue"])

    products = OrderItem.objects.filter(order__complete=True) \
        .values('product__name') \
        .annotate(
            total_sold=Sum('quantity'),
            revenue=Sum(
                ExpressionWrapper(
                    F('product__price') * F('quantity'),
                    output_field=DecimalField()
                )
            )
        )

    for p in products:
        writer.writerow([
            p['product__name'],
            p['total_sold'],
            float(p['revenue'] or 0)
        ])

    return response

def export_customers_excel(request):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers Report"

    # Header
    ws.append(["Customer", "Total Orders", "Total Spent"])

    # Header styling
    header_fill = PatternFill(start_color="0b2e66", end_color="0b2e66", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    customers = OrderItem.objects.filter(
        order__complete=True,
        order__customer__user__isnull=False
    ).values('order__customer__user__username') \
    .annotate(
        total_orders=Count('order', distinct=True),
        total_spent=Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        )
    ).order_by('-total_spent')

    for c in customers:
        ws.append([
            c['order__customer__user__username'],
            c['total_orders'],
            float(c['total_spent'] or 0)
        ])

    # Create table
    table = Table(displayName="CustomersTable", ref=f"A1:C{ws.max_row}")

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename="customers_report.xlsx"'

    wb.save(response)
    return response

def export_customers_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="customers_report.csv"'

    writer = csv.writer(response)
    writer.writerow(["Customer", "Total Orders", "Total Spent"])

    customers = OrderItem.objects.filter(order__complete=True) \
        .values('order__customer__user__username') \
        .annotate(
            total_orders=Count('order', distinct=True),
            total_spent=Sum(
                ExpressionWrapper(
                    F('product__price') * F('quantity'),
                    output_field=DecimalField()
                )
            )
        )

    for c in customers:
        writer.writerow([
            c['order__customer__user__username'],
            c['total_orders'],
            float(c['total_spent'] or 0)
        ])

    return response



def edit_order(request, order_id):
    order = Order.objects.get(id=order_id)

    if request.method == 'POST':
    #     # ✅ Update status
    #     complete_value = request.POST.get('complete')
    #     order.complete = True if complete_value == 'true' else False

        # ✅ Update order status
        order_status = request.POST.get('order_status')
        if order_status:
            order.order_status = order_status
        # ✅ Update date
        date_str = request.POST.get('date_ordered')
        if date_str:
            new_date = parse_datetime(date_str)
            order.date_ordered = new_date

        order.save()

        # ✅ ALSO UPDATE RELATED MODELS
        # 1. Invoice
        if hasattr(order, 'invoice'):
            order.invoice.issued_date = order.date_ordered
            order.invoice.save()

        # 2. Shipping Address
        shipping = ShippingAddress.objects.filter(order=order).first()
        if shipping:
            shipping.date_added = order.date_ordered
            shipping.save()

        return redirect('dashboard:orders')

    return render(request, 'dashboard/edit_order.html', {'order': order})


@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def delete_order(request, order_id):
    order = get_object_or_404(Order, id=order_id)

    # If admin clicks "Yes Delete"
    if request.method == "POST":
        order.delete()
        return redirect('dashboard:orders')

    # Show delete confirmation page
    return render(request, "dashboard/order_confirm_delete.html", {
        "order": order
    })

@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def edit_customer(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)

    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            return redirect('dashboard:dashboard_customers')
    else:
        form = CustomerForm(instance=customer)

    return render(request, 'dashboard/edit_customer.html', {'form': form})


@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def delete_customer(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)

    if request.method == "POST":
        customer.delete()
        return redirect('dashboard:dashboard_customers')

    return render(request, "dashboard/customer_confirm_delete.html", {
        "customer": customer
    })
@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def edit_order_summary(request, summary_id):
    summary = get_object_or_404(OrderSummary, id=summary_id)

    if request.method == 'POST':
        form = OrderSummaryForm(request.POST, instance=summary)
        if form.is_valid():
            form.save()
            return redirect('dashboard:dashboard_order_summaries')
    else:
        form = OrderSummaryForm(instance=summary)

    return render(request, 'dashboard/edit_order_summary.html', {'form': form,
        'summary': summary})


@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def edit_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)

    if request.method == 'POST':
        form = InvoiceForm(request.POST, instance=invoice)
        if form.is_valid():
            form.save()
            return redirect('dashboard:dashboard_invoices')
    else:
        form = InvoiceForm(instance=invoice)

    return render(request, 'dashboard/edit_invoice.html', {'form': form})

@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def delete_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)

    # When admin confirms deletion
    if request.method == "POST":
        invoice.delete()
        return redirect('dashboard:dashboard_invoices')

    # Show confirmation page
    return render(request, "dashboard/invoice_confirm_delete.html", {
        "invoice": invoice
    })

@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def edit_shipping_address(request, address_id):
    address = get_object_or_404(ShippingAddress, id=address_id)

    if request.method == 'POST':
        form = ShippingAddressForm(request.POST, instance=address)
        if form.is_valid():
            form.save()
            return redirect('dashboard:dashboard_shipping_addresses')
    else:
        form = ShippingAddressForm(instance=address)

    return render(request, 'dashboard/edit_shipping_address.html', {'form': form})

@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def delete_shipping_address(request, address_id):
    address = get_object_or_404(ShippingAddress, id=address_id)

    if request.method == "POST":
        address.delete()
        return redirect('dashboard:dashboard_shipping_addresses')

    return render(request, "dashboard/shipping_address_confirm_delete.html", {
        "address": address
    })

@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def edit_user(request, user_id):
    user_obj = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user_obj)
        if form.is_valid():
            form.save()
            return redirect('dashboard:users')
    else:
        form = UserEditForm(instance=user_obj)

    return render(request, 'dashboard/edit_user.html', {'form': form})

@login_required(login_url='/admin/login/')
@user_passes_test(is_admin)
def delete_user(request, user_id):
    user_obj = get_object_or_404(User, id=user_id)

    # Prevent deleting yourself
    if request.user == user_obj:
        messages.error(request, "You cannot delete your own account.")
        return redirect('dashboard:users')

    if request.method == "POST":
        user_obj.delete()
        return redirect('dashboard:users')

    return render(request, "dashboard/user_confirm_delete.html", {
        "user_obj": user_obj
    })

def feedback_list(request):
    feedbacks = Feedback.objects.all().order_by('-created_at')

    return render(request, 'dashboard/feedback.html', {
        'feedbacks': feedbacks
    })